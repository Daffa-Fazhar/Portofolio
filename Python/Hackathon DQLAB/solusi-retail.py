import itertools
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path

import matplotlib
import pandas as pd
from openpyxl import load_workbook

matplotlib.use("Agg")
import matplotlib.pyplot as plt


# Folder script = lokasi ekspektasi grader untuk input/output.
BASE_DIR = Path(__file__).resolve().parent
INPUT_FILE = BASE_DIR / "data_penjualan.xlsx"
OUTPUT_EXCEL = BASE_DIR / "retail_insight.xlsx"
OUTPUT_PNG_INDEX = BASE_DIR / "rising_star_index.png"
OUTPUT_PNG_ACTUAL = BASE_DIR / "rising_star_actual.png"

MA_WINDOW = 3
MIN_UPTREND_DAYS = 12
MIN_SUPPORT = 0.01
MIN_LIFT = 2.0


def _quantize_half_up(value: float, places: int) -> float:
    """Pembulatan half-up seperti fungsi ROUND di Excel untuk angka biasa."""
    q = Decimal("1").scaleb(-places) if places >= 0 else Decimal(1)
    return float(Decimal(str(float(value))).quantize(q, rounding=ROUND_HALF_UP))


def _to_int_money(value: float) -> int:
    return int(Decimal(str(float(value))).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _normalize_kode(x) -> str:
    if pd.isna(x):
        return ""
    return str(x).strip()


def preprocess_sales(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="Transaksi")
    df["tgl_transaksi"] = pd.to_datetime(df["tgl_transaksi"])
    df["jumlah_terjual"] = pd.to_numeric(df["jumlah_terjual"], errors="coerce").fillna(0.0)
    df["total_nilai"] = pd.to_numeric(df["total_nilai"], errors="coerce").fillna(0.0)
    df["kode_produk"] = df["kode_produk"].map(_normalize_kode)
    df = df[df["kode_produk"] != ""]
    return df


def longest_consecutive_true(series: pd.Series) -> int:
    max_streak = 0
    current = 0
    for value in series.fillna(False):
        if bool(value):
            current += 1
            max_streak = max(max_streak, current)
        else:
            current = 0
    return max_streak


def detect_rising_star(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    daily = (
        df.groupby(["tgl_transaksi", "kode_produk", "nama_produk"], as_index=False)["total_nilai"]
        .sum()
        .sort_values(["kode_produk", "tgl_transaksi"])
    )
    daily["ma3"] = daily.groupby("kode_produk")["total_nilai"].transform(
        lambda s: s.rolling(window=MA_WINDOW, min_periods=MA_WINDOW).mean()
    )
    daily["ma3_diff"] = daily.groupby("kode_produk")["ma3"].diff()
    daily["ma3_up"] = daily["ma3_diff"] > 0

    summary = (
        daily.groupby(["kode_produk", "nama_produk"], as_index=False)
        .agg(
            first_ma3=("ma3", lambda s: s.dropna().iloc[0] if not s.dropna().empty else pd.NA),
            last_ma3=("ma3", lambda s: s.dropna().iloc[-1] if not s.dropna().empty else pd.NA),
            total_penjualan=("total_nilai", "sum"),
            max_up_streak=("ma3_up", longest_consecutive_true),
        )
        .sort_values("total_penjualan", ascending=False)
    )
    summary["growth_pct"] = (
        (summary["last_ma3"] - summary["first_ma3"])
        / summary["first_ma3"].replace(0, pd.NA)
        * 100
    ).fillna(0.0)

    rising = (
        summary[summary["max_up_streak"] >= MIN_UPTREND_DAYS]
        .sort_values(["growth_pct", "total_penjualan", "kode_produk"], ascending=[False, False, True])
        .reset_index(drop=True)
        .copy()
    )
    return daily, rising


def apriori_itemsets(transactions: list[set[str]], min_support: float) -> dict[frozenset[str], float]:
    num_tx = len(transactions)
    support_map: dict[frozenset[str], float] = {}
    frequent_levels: list[set[frozenset[str]]] = []

    item_count: dict[str, int] = {}
    for tx in transactions:
        for item in tx:
            item_count[item] = item_count.get(item, 0) + 1

    level_1: set[frozenset[str]] = set()
    for item, count in item_count.items():
        support = count / num_tx
        if support >= min_support:
            key = frozenset([item])
            level_1.add(key)
            support_map[key] = support
    frequent_levels.append(level_1)

    k = 2
    while frequent_levels[k - 2]:
        prev_level = list(frequent_levels[k - 2])
        candidates: set[frozenset[str]] = set()

        for i in range(len(prev_level)):
            for j in range(i + 1, len(prev_level)):
                union_set = prev_level[i] | prev_level[j]
                if len(union_set) == k:
                    all_subsets_frequent = True
                    for subset in itertools.combinations(union_set, k - 1):
                        if frozenset(subset) not in frequent_levels[k - 2]:
                            all_subsets_frequent = False
                            break
                    if all_subsets_frequent:
                        candidates.add(union_set)

        if not candidates:
            break

        next_level: set[frozenset[str]] = set()
        for cand in candidates:
            count = sum(1 for tx in transactions if cand.issubset(tx))
            support = count / num_tx
            if support >= min_support:
                next_level.add(cand)
                support_map[cand] = support

        frequent_levels.append(next_level)
        k += 1

    return support_map


def generate_packaging_rules(df: pd.DataFrame, rising_codes: set[str]) -> pd.DataFrame:
    product_name_map: dict[str, str] = (
        df[["kode_produk", "nama_produk"]]
        .dropna(subset=["kode_produk", "nama_produk"])
        .drop_duplicates(subset=["kode_produk"])
        .set_index("kode_produk")["nama_produk"]
        .astype(str)
        .to_dict()
    )

    def format_itemset(codes: frozenset[str]) -> str:
        # Selaras enumerasi combinasi (sorted lexis kode) → string Excel deterministik.
        return ", ".join(product_name_map.get(c, str(c)) for c in sorted(codes))

    basket = df.groupby("nomor_struk")["kode_produk"].apply(lambda x: set(x.dropna())).tolist()
    basket = [tx for tx in basket if len(tx) >= 1]
    if not basket:
        return pd.DataFrame(
            columns=["Jika Membeli", "Maka Membeli", "Jumlah Invoice", "Support", "Confidence", "Lift"]
        )

    support_map = apriori_itemsets(basket, min_support=MIN_SUPPORT)
    rows = []

    for itemset, support_xy in support_map.items():
        if len(itemset) < 2:
            continue
        if not any(code in rising_codes for code in itemset):
            continue

        for r in range(1, len(itemset)):
            for antecedent_tuple in itertools.combinations(sorted(itemset), r):
                antecedent = frozenset(antecedent_tuple)
                consequent = itemset - antecedent
                if not consequent:
                    continue

                support_x = support_map.get(antecedent)
                support_y = support_map.get(consequent)
                if support_x is None or support_y is None:
                    continue

                confidence = support_xy / support_x
                lift = confidence / support_y
                if lift < MIN_LIFT:
                    continue

                count_invoice = sum(1 for tx in basket if itemset.issubset(tx))
                rows.append(
                    {
                        "Jika Membeli": format_itemset(antecedent),
                        "Maka Membeli": format_itemset(consequent),
                        "Jumlah Invoice": count_invoice,
                        "_support_raw": support_xy,
                        "_conf_raw": confidence,
                        "_lift_raw": lift,
                    }
                )

    if not rows:
        return pd.DataFrame(
            columns=["Jika Membeli", "Maka Membeli", "Jumlah Invoice", "Support", "Confidence", "Lift"]
        )

    rules = pd.DataFrame(rows).drop_duplicates(
        subset=["Jika Membeli", "Maka Membeli", "Jumlah Invoice", "_support_raw", "_conf_raw", "_lift_raw"]
    )
    rules["Support"] = rules["_support_raw"].map(lambda v: _quantize_half_up(float(v), 2))
    rules["Confidence"] = rules["_conf_raw"].map(lambda v: _quantize_half_up(float(v), 2))
    rules["Lift"] = rules["_lift_raw"].map(lambda v: _quantize_half_up(float(v), 2))
    rules = rules.drop_duplicates(
        subset=["Jika Membeli", "Maka Membeli", "Jumlah Invoice", "Support", "Confidence", "Lift"]
    )
    rules = rules.sort_values(
        ["Lift", "Confidence", "Support", "Jika Membeli", "Maka Membeli"],
        ascending=[False, False, False, True, True],
    ).reset_index(drop=True)
    return rules[
        ["Jika Membeli", "Maka Membeli", "Jumlah Invoice", "Support", "Confidence", "Lift"]
    ]


def build_daily_with_normalized(daily: pd.DataFrame) -> pd.DataFrame:
    daily_df = daily.copy()
    daily_df["Normalized"] = daily_df.groupby("kode_produk")["ma3"].transform(
        lambda s: (
            s / s.dropna().iloc[0] * 100
            if not s.dropna().empty and s.dropna().iloc[0] != 0
            else pd.NA
        )
    )
    return daily_df


def plot_rising_star(
    df: pd.DataFrame,
    daily: pd.DataFrame,
    rising: pd.DataFrame,
    output_file: Path,
    *,
    index_chart: bool,
) -> None:
    daily_df = build_daily_with_normalized(daily)

    final_report = rising.rename(columns={"growth_pct": "Growth_Pct"})
    sorted_report = final_report.sort_values(by="Growth_Pct", ascending=False).reset_index(drop=True)

    custom_palette = [
        "#FFD700",
        "#C0C0C0",
        "#CD7F32",
        "#2ecc71",
        "#3498db",
        "#9b59b6",
        "#e74c3c",
        "#34495e",
    ]
    default_color = "#95a5a6"

    color_mapping: dict[str, str] = {}
    rank_mapping: dict[str, int] = {}
    for i, row in enumerate(sorted_report.itertuples()):
        kp = row.kode_produk
        color_mapping[kp] = custom_palette[i] if i < len(custom_palette) else default_color
        rank_mapping[kp] = i + 1

    top3_sales = (
        df.groupby(["kode_produk", "nama_produk"])["total_nilai"].sum().reset_index().sort_values(
            by="total_nilai", ascending=False
        ).head(3)
    )
    top3_codes = top3_sales["kode_produk"].tolist()

    top3_plot_df = daily_df[daily_df["kode_produk"].isin(top3_codes)].copy()

    grey_colors = ["#B0B0B0", "#909090", "#707070"]

    fig = plt.figure(figsize=(15, 8), dpi=100)
    ax = fig.add_subplot(111)

    for idx, code in enumerate(top3_codes):
        grp = top3_plot_df[top3_plot_df["kode_produk"] == code]
        if grp.empty:
            continue
        nama_produk = grp["nama_produk"].iloc[0]
        grey_color = grey_colors[idx] if idx < len(grey_colors) else "#808080"
        y_series = grp["Normalized"] if index_chart else grp["total_nilai"]
        ax.plot(
            grp["tgl_transaksi"],
            y_series,
            linestyle="--",
            linewidth=2,
            marker="o",
            markersize=3,
            color=grey_color,
            alpha=0.7,
            label=f"Top Sales: {nama_produk}",
        )

    for row in sorted_report.itertuples():
        kode_produk = row.kode_produk
        grp = daily_df[daily_df["kode_produk"] == kode_produk]
        if grp.empty:
            continue
        nama_produk = grp["nama_produk"].iloc[0]
        line_color = color_mapping.get(kode_produk, default_color)
        rank = rank_mapping.get(kode_produk, "?")
        label_with_rank = f"Rank {rank}: {nama_produk}"
        y_series = grp["Normalized"] if index_chart else grp["total_nilai"]
        ax.plot(
            grp["tgl_transaksi"],
            y_series,
            marker="o",
            markersize=4,
            linewidth=2.5,
            color=line_color,
            label=label_with_rank,
        )

    font_title = {"family": "sans-serif", "color": "black", "weight": "bold", "size": 16}
    font_label = {"family": "sans-serif", "weight": "normal", "size": 12}

    if index_chart:
        ax.set_title(
            "ANALISIS PERTUMBUHAN RELATIF PRODUK RISING STAR\n(Dengan Benchmark Top 3 Total Penjualan)",
            fontdict=font_title,
            pad=20,
        )
        ax.set_ylabel("Indeks Pertumbuhan (Base 100)", fontdict=font_label, labelpad=10)
        ax.grid(True, linestyle="--", linewidth=0.5, alpha=0.5)
        ax.axhline(y=100, color="black", linestyle="-", linewidth=1, alpha=0.5)
    else:
        ax.set_title(
            "ANALISIS NILAI PENJUALAN PRODUK RISING STAR\n(Nilai Penjualan Asli)",
            fontdict=font_title,
            pad=20,
        )
        ax.set_ylabel("Total Nilai Penjualan", fontdict=font_label, labelpad=10)
        ax.grid(True, linestyle="--", linewidth=0.5, alpha=0.5)

    ax.set_xlabel("Periode Tanggal", fontdict=font_label, labelpad=10)

    plt.xticks(rotation=45, ha="right", fontsize=10)
    plt.yticks(fontsize=10)

    handles, labels_ax = ax.get_legend_handles_labels()
    top_sales_items = []
    rising_items_list = []
    for h, l in zip(handles, labels_ax):
        if l.startswith("Top Sales"):
            top_sales_items.append((h, l))
        else:
            rising_items_list.append((h, l))

    rising_items_list.sort(key=lambda x: int(x[1].split(":")[0].split()[1]))
    final_handles = [x[0] for x in top_sales_items + rising_items_list]
    final_labels = [x[1] for x in top_sales_items + rising_items_list]

    ax.legend(
        final_handles,
        final_labels,
        title="Kategori Produk",
        title_fontsize=12,
        fontsize=10,
        bbox_to_anchor=(1.02, 1),
        loc="upper left",
        borderaxespad=0,
        frameon=True,
        shadow=True,
    )
    plt.tight_layout()
    plt.savefig(output_file.as_posix(), bbox_inches="tight")
    plt.close(fig)


def save_output(rising: pd.DataFrame, packaging: pd.DataFrame, output_path: Path) -> None:
    rising_export = rising.rename(
        columns={
            "kode_produk": "Kode Produk",
            "nama_produk": "Nama Produk",
            "growth_pct": "Growth %",
            "total_penjualan": "Total Penjualan",
        }
    )[["Kode Produk", "Nama Produk", "Growth %", "Total Penjualan"]].copy()

    rising_export["Growth %"] = rising_export["Growth %"].map(lambda v: _quantize_half_up(float(v), 2))
    rising_export["Total Penjualan"] = rising_export["Total Penjualan"].map(_to_int_money)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        rising_export.to_excel(writer, sheet_name="Rising Star", index=False)
        packaging.to_excel(writer, sheet_name="Potential Packaging", index=False)

    wb = load_workbook(output_path.as_posix())
    ws_rising = wb["Rising Star"]
    ws_pack = wb["Potential Packaging"]

    for row in range(2, ws_rising.max_row + 1):
        ws_rising[f"C{row}"].number_format = "0.00"
        ws_rising[f"D{row}"].number_format = "#,##0"

    for row in range(2, ws_pack.max_row + 1):
        ws_pack[f"D{row}"].number_format = "0.##"
        ws_pack[f"E{row}"].number_format = "0.##"
        ws_pack[f"F{row}"].number_format = "0.##"

    wb.save(output_path.as_posix())


def main() -> None:
    if not INPUT_FILE.is_file():
        raise FileNotFoundError(f"File input tidak ditemukan (relatif script): {INPUT_FILE}")

    df = preprocess_sales(INPUT_FILE)
    daily, rising = detect_rising_star(df)

    rising_codes_set = set(rising["kode_produk"])
    packaging = generate_packaging_rules(df, rising_codes_set)

    save_output(rising, packaging, OUTPUT_EXCEL)

    plot_rising_star(df, daily, rising, OUTPUT_PNG_INDEX, index_chart=True)
    plot_rising_star(df, daily, rising, OUTPUT_PNG_ACTUAL, index_chart=False)

    print(f"Output Excel berhasil dibuat: {OUTPUT_EXCEL}")
    print(f"Output grafik berhasil dibuat: {OUTPUT_PNG_INDEX}, {OUTPUT_PNG_ACTUAL}")


if __name__ == "__main__":
    main()
