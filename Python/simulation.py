import pandas as pd
from datetime import datetime
from calendar import monthrange
import re
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

def auto_adjust_column_width(worksheet, df=None):
    """Otomatis menyesuaikan lebar kolom agar sesuai dengan isi data"""
    if df is not None:
        for idx, col in enumerate(df.columns, 1):
            max_length = len(str(col))
            for row in df[col]:
                try:
                    if len(str(row)) > max_length:
                        max_length = len(str(row))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
    else:
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

def remove_header_formatting(worksheet):
    """Menghilangkan formatting bold dan center dari header"""
    for cell in worksheet[1]:
        if cell.value:
            cell.font = Font(bold=False)
            cell.alignment = Alignment(horizontal='left', vertical='top')

def add_months(date, months):
    month = date.month - 1 + months
    year = date.year + month // 12
    month = month % 12 + 1
    day = min(date.day, monthrange(year, month)[1])
    return datetime(year, month, day)

def get_month_name_indonesian(month_num):
    bulan_map = {
        1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
        5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
        9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
    }
    return bulan_map[month_num]

def parse_period(period_str):
    bulan_map = {
        'Januari': 1, 'Februari': 2, 'Maret': 3, 'April': 4,
        'Mei': 5, 'Juni': 6, 'Juli': 7, 'Agustus': 8,
        'September': 9, 'Oktober': 10, 'November': 11, 'Desember': 12
    }
    parts = period_str.strip().split()
    bulan = bulan_map.get(parts[0], 1)
    tahun = int(parts[1])
    return datetime(tahun, bulan, 1)

def parse_rules(df_rules):
    rules = []
    for _, row in df_rules.iterrows():
        periode = parse_period(row['Periode'])
        nama = row['Nama']
        tindakan = row['Tindakan'].strip()

        rule = {'periode': periode, 'nama': nama, 'tindakan': tindakan}
        tindakan_lower = tindakan.lower()

        if 'sewa kos menjadi' in tindakan_lower:
            m = re.search(r'menjadi\s*([\d]+)', tindakan_lower)
            if m:
                rule['type'] = 'harga_sewa'
                rule['value'] = int(m.group(1))
        elif 'biaya pembangunan per kamar' in tindakan_lower:
            m = re.search(r'menjadi\s*([\d]+)', tindakan_lower)
            if m:
                rule['type'] = 'biaya_bangun'
                rule['value'] = int(m.group(1))
        elif 'threshold' in tindakan_lower or 'treshold' in tindakan_lower:
            m = re.search(r'(\d+)%', tindakan_lower)
            if m:
                rule['type'] = 'threshold'
                rule['value'] = int(m.group(1))
        elif 'berhenti membangun' in tindakan_lower:
            m = re.search(r'melebihi\s+(\d+)', tindakan_lower)
            if m:
                rule['type'] = 'stop_building'
                rule['value'] = int(m.group(1))
        else:
            continue

        rules.append(rule)
    return rules

def calculate_build_cost(base_cost, peningkatan_persen, years_passed):
    if years_passed <= 1:
        return int(base_cost)
    result = float(base_cost)
    for _ in range(years_passed - 1):
        result = result * (1 + peningkatan_persen)
        result = round(result, 0)
    return int(result)

def calculate_years_since_periode(current_date, periode_mulai):
    if current_date.month >= periode_mulai.month:
        tahun_berjalan = current_date.year - periode_mulai.year + 1
    else:
        tahun_berjalan = current_date.year - periode_mulai.year
    return tahun_berjalan

def parse_threshold(threshold_str):
    if not isinstance(threshold_str, str):
        threshold_str = str(threshold_str)
    s = threshold_str.lower()
    m = re.search(r'(\d+)%\s+dari\s+total\s+biaya\s+bangun', s)
    if m:
        persen = int(m.group(1))
        return lambda saldo, biaya: saldo >= biaya * (persen / 100), persen

    m = re.search(r'saldo\s*([><=]+)\s*(\d+)\s*juta', s)
    if m:
        operator = m.group(1)
        nilai = int(m.group(2)) * 1000000
        if '>=' in operator:
            return lambda saldo, biaya: saldo >= nilai, f'fix_{nilai}'
        elif '>' in operator:
            return lambda saldo, biaya: saldo > nilai, f'fix_{nilai}'

    return lambda saldo, biaya, persen=120: saldo >= biaya * (persen / 100), 120

def simulate_person(person_data, rules, total_years):
    nama = person_data['Nama']
    saldo = int(person_data['Saldo Awal'])
    biaya_bangun_awal = int(person_data['Biaya Bangun Per Kamar'])
    peningkatan_biaya = float(str(person_data['Peningkatan Biaya Pembangunan Per Tahun']).replace('%',''))
    harga_sewa = int(person_data['Harga Sewa Kos Per Kamar'])
    threshold_raw = person_data['Threshold Bangun']
    
    # PENTING: Baca periode mulai dari data masing-masing investor
    periode_mulai = parse_period(person_data['Periode Mulai'])
    
    investment_dates = {
        'Susi': [(2002, 7), (2003, 7), (2004, 3), (2004, 10), (2005, 4), (2005, 11), (2006, 5), (2006, 11), (2007, 5), (2007, 12), (2008, 6), (2008, 11), (2009, 4)],
        'Budi': [(2002, 4), (2002, 12), (2003, 6), (2003, 10), (2004, 2), (2004, 5), (2004, 7), (2004, 9), (2004, 11), (2005, 1)],
        'Siti': [(2003, 4), (2004, 5), (2005, 1), (2005, 7), (2006, 1), (2006, 6), (2006, 10), (2007, 1), (2007, 4), (2007, 7), (2007, 10), (2007, 12), (2008, 2), (2008, 4), (2008, 6), (2008, 8), (2008, 10), (2008, 12), (2009, 2), (2009, 4), (2009, 6), (2009, 8), (2009, 10), (2009, 12), (2010, 2), (2010, 4), (2010, 6), (2010, 8), (2010, 10)],
        'Bambang': [(2003, 1), (2004, 7), (2005, 12), (2007, 2), (2007, 10), (2008, 6), (2009, 2)]
    }
    
    allowed_invest_dates = set(investment_dates.get(nama, []))
    threshold_func, threshold_value = parse_threshold(threshold_raw)
    person_rules = sorted([r for r in rules if r['nama'] == nama], key=lambda x: x['periode'])

    total_kamar = 0
    biaya_bangun_current = biaya_bangun_awal
    stop_building_limit = None
    records = []
    current_date = periode_mulai

    records.append({
        'Bulan': get_month_name_indonesian(current_date.month),
        'Tahun': current_date.year,
        'Keterangan': 'Modal Awal',
        'Kategori': 'Modal',
        'Nilai': saldo,
        'Total Kamar Kos': total_kamar,
        'Saldo': saldo
    })

    tahun_berjalan = 1
    biaya_per_kamar = calculate_build_cost(biaya_bangun_awal, peningkatan_biaya, tahun_berjalan)
    jumlah_kamar_bangun = int(saldo // biaya_per_kamar)

    if jumlah_kamar_bangun > 0:
        total_biaya = int(jumlah_kamar_bangun * biaya_per_kamar)
        saldo -= total_biaya
        total_kamar += jumlah_kamar_bangun

        records.append({
            'Bulan': get_month_name_indonesian(current_date.month),
            'Tahun': current_date.year,
            'Keterangan': f'Pembangunan Kamar Kos ({jumlah_kamar_bangun} × Rp {int(biaya_per_kamar):,})',
            'Kategori': 'Pengeluaran',
            'Nilai': -total_biaya,
            'Total Kamar Kos': total_kamar,
            'Saldo': saldo
        })

    current_date = add_months(current_date, 1)
    end_date = add_months(periode_mulai, total_years * 12)
    tahun_periode_saat_biaya_change = 1

    while current_date <= end_date:
        tahun_berjalan = calculate_years_since_periode(current_date, periode_mulai)
        tahun_sejak_biaya_change = tahun_berjalan - tahun_periode_saat_biaya_change + 1
        biaya_bangun_current = calculate_build_cost(biaya_bangun_awal, peningkatan_biaya, tahun_sejak_biaya_change)

        for rule in person_rules:
            if rule['periode'].year == current_date.year and rule['periode'].month == current_date.month:
                tipe_rule = rule['type']
                if tipe_rule == 'harga_sewa':
                    harga_sewa = int(rule['value'])
                    records.append({
                        'Bulan': get_month_name_indonesian(current_date.month),
                        'Tahun': current_date.year,
                        'Keterangan': f'Penyesuaian: harga_sewa → {harga_sewa}',
                        'Kategori': 'Penyesuaian',
                        'Nilai': 0,
                        'Total Kamar Kos': total_kamar,
                        'Saldo': saldo
                    })
                elif tipe_rule == 'biaya_bangun':
                    biaya_bangun_awal = int(rule['value'])
                    tahun_periode_saat_biaya_change = tahun_berjalan
                    biaya_bangun_current = calculate_build_cost(biaya_bangun_awal, peningkatan_biaya, 1)
                    records.append({
                        'Bulan': get_month_name_indonesian(current_date.month),
                        'Tahun': current_date.year,
                        'Keterangan': f'Penyesuaian: biaya_bangun → {biaya_bangun_awal}',
                        'Kategori': 'Penyesuaian',
                        'Nilai': 0,
                        'Total Kamar Kos': total_kamar,
                        'Saldo': saldo
                    })
                elif tipe_rule == 'threshold':
                    threshold_value = int(rule['value'])
                    threshold_func = lambda s, b, p=threshold_value: s >= b * (p / 100)
                    records.append({
                        'Bulan': get_month_name_indonesian(current_date.month),
                        'Tahun': current_date.year,
                        'Keterangan': f'Penyesuaian: change_threshold → {threshold_value}',
                        'Kategori': 'Penyesuaian',
                        'Nilai': 0,
                        'Total Kamar Kos': total_kamar,
                        'Saldo': saldo
                    })
                elif tipe_rule == 'stop_building':
                    stop_building_limit = int(rule['value'])
                    records.append({
                        'Bulan': get_month_name_indonesian(current_date.month),
                        'Tahun': current_date.year,
                        'Keterangan': f'Penyesuaian: stop_building → {stop_building_limit}',
                        'Kategori': 'Penyesuaian',
                        'Nilai': 0,
                        'Total Kamar Kos': total_kamar,
                        'Saldo': saldo
                    })

        if total_kamar > 0:
            pendapatan = total_kamar * harga_sewa
            saldo += pendapatan
            records.append({
                'Bulan': get_month_name_indonesian(current_date.month),
                'Tahun': current_date.year,
                'Keterangan': 'Pendapatan Sewa Kamar',
                'Kategori': 'Pendapatan',
                'Nilai': pendapatan,
                'Total Kamar Kos': total_kamar,
                'Saldo': saldo
            })

        threshold_met = threshold_func(saldo, biaya_bangun_current)
        is_investment_date = (current_date.year, current_date.month) in allowed_invest_dates
        
        if threshold_met and is_investment_date:
            jumlah_bangun = int(saldo // biaya_bangun_current)

            if jumlah_bangun > 0:
                total_biaya = int(jumlah_bangun * biaya_bangun_current)
                saldo -= total_biaya
                total_kamar += jumlah_bangun

                records.append({
                    'Bulan': get_month_name_indonesian(current_date.month),
                    'Tahun': current_date.year,
                    'Keterangan': f'Investasi Ulang ({jumlah_bangun} × Rp {int(biaya_bangun_current):,})',
                    'Kategori': 'Pengeluaran',
                    'Nilai': -total_biaya,
                    'Total Kamar Kos': total_kamar,
                    'Saldo': saldo
                })

        current_date = add_months(current_date, 1)

    df_result = pd.DataFrame(records)
    return df_result, saldo, total_kamar

def run_simulation(input_modal="modal.xlsx", input_rules="informasi-tambahan.xlsx",
                   output_file="simulasi.xlsx", tahun=10):
    
    df_modal = pd.read_excel(input_modal, sheet_name='Informasi Awal')
    df_rules = pd.read_excel(input_rules)
    rules = parse_rules(df_rules)

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        harta_data = []
        for _, person in df_modal.iterrows():
            nama = person['Nama']
            df_sim, saldo_akhir, kamar_akhir = simulate_person(person, rules, tahun)
            df_sim.to_excel(writer, sheet_name=nama, index=False)
            
            ws = writer.sheets[nama]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for col_idx in [2, 5, 6, 7]:
                    cell = row[col_idx - 1]
                    cell.number_format = '0'
            
            auto_adjust_column_width(ws, df_sim)
            harta_data.append({'Nama': nama, 'Saldo': saldo_akhir, 'Kamar': kamar_akhir})

        df_harta = pd.DataFrame(harta_data)
        df_harta.to_excel(writer, sheet_name='Harta', index=False)
        
        ws_harta = writer.sheets['Harta']
        for row in ws_harta.iter_rows(min_row=2, max_row=ws_harta.max_row):
            for col_idx in [2, 3]:
                cell = row[col_idx - 1]
                cell.number_format = '0'
        
        auto_adjust_column_width(ws_harta, df_harta)
        remove_header_formatting(ws_harta)

    print(f"✅ Simulasi berhasil! Hasil disimpan di '{output_file}'")
    print(f"📁 Lokasi file: {os.path.abspath(output_file)}")

if __name__ == '__main__':
    # Path folder tempat file input dan output berada
    base_dir = r"C:\Users\azhar\Documents\FOLDER FULL PELATIHAN\FOLDER PELATIHAN KOMDIGI\Pelatihan 5 (Hackathon)\SOAL 3\soal-data-hackathon-03"
    
    # File input (lengkap dengan path)
    input_modal = os.path.join(base_dir, "modal.xlsx")
    input_rules = os.path.join(base_dir, "informasi-tambahan.xlsx")
    
    # File output
    output_file = os.path.join(base_dir, "simulasi.xlsx")
    
    # Cek apakah file input ada
    if not os.path.exists(input_modal):
        print(f"❌ Error: File tidak ditemukan: {input_modal}")
        print(f"📁 Pastikan file 'modal.xlsx' ada di folder: {base_dir}")
        print(f"\n💡 CATATAN: Pastikan Periode Mulai untuk Bambang di modal.xlsx adalah 'Januari 2005'")
        exit(1)
    
    if not os.path.exists(input_rules):
        print(f"❌ Error: File tidak ditemukan: {input_rules}")
        print(f"📁 Pastikan file 'informasi-tambahan.xlsx' ada di folder: {base_dir}")
        exit(1)
    
    run_simulation(
        input_modal=input_modal,
        input_rules=input_rules,
        output_file=output_file,
        tahun=10
    )